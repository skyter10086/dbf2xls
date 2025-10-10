# file: no-sweat.py

import xlwings as xw
import pandas as pd
from dbfread2 import DBF
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
import loguru
import typer

# *** 公共参数 ***
current_term = "202510"
source_base  = r"D:\企业补贴\数据"
dest_base = r"D:\企业补贴\银行报盘"
bank_templates = {
    "工行":{
        "temp_path":r"D:\企业补贴\银行报盘\工商银行报盘模板.xlsx",
        "sheet":"工行跨行",
        "cell":"A2",
    },
    "建行":{
        "temp_path":r"D:\企业补贴\银行报盘\建设银行报盘模板.xlsx",
        "sheet":"sheet1",
        "cell":"A2",
    },
    "中行-油区":{
        "temp_path":r"D:\企业补贴\银行报盘\中国银行报盘模板.xlsx",
        "sheet":"sheet1",
        "cell":"A2",
    },
    "中行-南阳":{
        "temp_path": r"D:\企业补贴\银行报盘\中国银行南阳报盘模板.xlsx",
        "sheet":"sheet1",
        "cell":"A2",
    }
}


summary_templates = {
    "老人": {
        "path": r"D:\企业补贴\银行报盘\离退休职工企业补贴汇总模板（老人）.xlsx",
        "sheets":[
            {
                "name": "单位汇总",
                "title": {
                    "cell":"A1",
                    "value":f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"全民离退休人员企业补贴发放汇总表(单位)"}"
                },
                "grid":{
                    "arch": "B4",
                }
            },
            {
                "name": "银行汇总",
                "title": {
                    "cell":"A1",
                    "value":f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"全民离退休人员企业补贴发放汇总表(单位)"}"
                },
                "grid":{
                    "arch": "K4",
                }
            },
        ]
    },
 #   "中人": {
  #      "path":,
   #     "sheet"
    #},
    "集体工": {
        "path":r"D:\企业补贴\银行报盘\退休集体工企业补贴汇总模板.xlsx",
        "sheets": [
            {
                "name": "企业补贴（财务拨款）",
                "title": {
                    "cell": "A1",
                    "value": f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"退休集体工企业补贴发放汇总表"}",
                },
                "grid":{
                    "arch": "B4",
                }
            },
                        {
                "name": "企业补贴（含遗孀）",
                "title": {
                    "cell": "A1",
                    "value": f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(含遗孀)"}",
                },
                "grid":{
                    "arch": "B4",
                }
            },
                        {
                "name": "企业补贴（单位征集）",
                "title": {
                    "cell": "A1",
                    "value": f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(工程)"}",
                },
                "grid":{
                    "arch": "L4",
                }
            },
                        {
                "name": "企业补贴（银行汇总）",
                "title": {
                    "cell": "A1",
                    "value": f"{current_term[:4]+"年"+str(int(current_term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(银行)"}",
                },
                "grid":{
                    "arch": "J4",
                }
            },
        ]
    },
}


# *** 根据模板导出数据 ***
def source_path(base, term):
    path = {
        "老人": Path(base) / f"{"企业补贴"+term}",
        "中人": Path(base) / f"{"提高待遇"+term}",
        "集体工": Path(base) / f"{"集体工企业补贴"+term}"
    }
    return path

def dest_path(base,term):
    path = Path(base) / f"{term[:4]+"年"}" / f"{str(int(term[-2:]))+"月"}"
    paths = {"工行":path / "工行报盘", "建行":path / "建行报盘", "中行":path / "中行报盘"}
    for _,p in paths.items():
        if not p.exists():
            p.mkdir(parents=True,exist_ok=True)
    return paths

def read_dbf(file_path:Path):
    if not file_path.exists():
        return None
    
    table = DBF(file_path,lowercase_names=True)
    data = pd.DataFrame(iter(table))
    
    data["补贴更正"] = pd.to_numeric(data["补贴更正"], errors="coerce").fillna(0)
    data["误餐补贴"] = pd.to_numeric(data["误餐补贴"], errors="coerce").fillna(0)
    data["补发补贴"] = pd.to_numeric(data["补发补贴"], errors="coerce").fillna(0)
    data["扣款_补贴"] = pd.to_numeric(data["扣款_补贴"], errors="coerce").fillna(0)
    data["补发_其它"] = pd.to_numeric(data["补发_其它"], errors="coerce").fillna(0)
    data["其它扣款"] = pd.to_numeric(data["其它扣款"], errors="coerce").fillna(0)
    data["实发补贴"] = pd.to_numeric(data["实发补贴"], errors="coerce").fillna(0)
    data["应发补贴"] = pd.to_numeric(data["应发补贴"], errors="coerce").fillna(0)
    return data

def add_index(df:pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return None
    df = df.reset_index(drop=True)
    df.index = df.index + 1
    df = df.reset_index()
    return df

def preprocess(df_,benifits_type):
    if  df_ is None:
        return None

    df = df_.loc[df_['re']==1,
        ["补贴更正","误餐补贴","补发补贴","扣款_补贴","补发_其它","其它扣款","姓名","身份证",
         "x_银行帐号","发放银行","银行帐号","收款行行号","发放地点","实发补贴","应发补贴","dwbm"]
    ]

    match benifits_type:
        case "老人企业补贴":
            df["企业补贴"] = df["补贴更正"] + df["误餐补贴"] + df["补发补贴"] - df["扣款_补贴"]
            df["提高待遇"] = df["补发_其它"] - df["其它扣款"]
            df["企业补贴"] = pd.to_numeric(df["企业补贴"], errors="coerce").fillna(0)
            df["提高待遇"] = pd.to_numeric(df["提高待遇"], errors="coerce").fillna(0)
            df.loc[df['发放银行']=='工行异地', '发放银行'] = '工商银行_跨行'
            df["账号地址"] = df["发放地点"]
            df["跨行行号"] = df["银行帐号"]
            return df

        case "集体工企业补贴":
            df["企业补贴"] = df["补贴更正"] + df["误餐补贴"] + df["补发补贴"] - df["扣款_补贴"]
            df["提高待遇"] = df["补发_其它"] - df["其它扣款"]
            df["企业补贴"] = pd.to_numeric(df["企业补贴"], errors="coerce").fillna(0)
            df["提高待遇"] = pd.to_numeric(df["提高待遇"], errors="coerce").fillna(0)
            df.loc[(df['发放银行']=='工商银行异地') | (df['发放银行'] == '商业银行（工行代发）'), '发放银行'] = '工商银行_跨行'
            df["账号地址"] = df["发放地点"]
            df["跨行行号"] = df["银行帐号"]
            return df

        case "中人提高待遇":
            df['提高待遇'] = df["补贴更正"] + df["补发补贴"] - df["其它扣款"]
            df['企业补贴'] = 0
            df["提高待遇"] = pd.to_numeric(df["提高待遇"], errors="coerce").fillna(0)
            df.loc[(df['发放银行']=='工商银行（异地）') & (df['发放银行'] == '交通银行'), '发放银行'] = '工商银行_跨行'
            df["账号地址"] = df["发放地点"]
            df["跨行行号"] = df["收款行行号"]
            return df
        
        case _:
            pass


def conv_icbc(df):
    if df is None:
        return None
    # 跨行                  
    df_1 = df.loc[df["发放银行"]=="工商银行_跨行"].copy() 
    df_1["行别"] = 1
    df_1["业务种类"] = "00602"
    df_1["协议书号"] = ""
    df_1["账号地址"] = df_1["发放地点"]
    df_1["跨行行号"] = df_1["银行帐号"]

    if not df_1.query("提高待遇 > 0").empty:
        result_1 = pd.concat([
            df_1.loc[df_1["企业补贴"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","企业补贴"]].copy(),
            df_1.loc[
                df_1["提高待遇"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","提高待遇"]].copy()
                .rename(columns={"提高待遇":"企业补贴"})
        ])
    else:
        result_1 = df_1.loc[df_1["企业补贴"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","企业补贴"]].copy()

    # 本行        
    df_2 = df.loc[df["发放银行"]=="工商银行"].copy() 
    df_2["行别"] = ""
    df_2["业务种类"] = ""
    df_2["协议书号"] = ""
    df_2["账号地址"] = ""
    df_2["跨行行号"] = ""
            
    if not df_2.query("提高待遇 > 0").empty:
        result_2 = pd.concat([
            df_2.loc[df_2["企业补贴"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","企业补贴"]].copy(),
            df_2.loc[
                df_2["提高待遇"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","提高待遇"]].copy()
                .rename(columns={"提高待遇":"企业补贴"})
        ])
    else:
        result_2 = df_2.loc[df_2["企业补贴"]>0,["姓名","x_银行帐号","行别","跨行行号","业务种类","协议书号","账号地址","企业补贴"]].copy()
    # 合并        
    result = pd.concat([result_1,result_2],ignore_index=True)
    if result.empty:
        return None
    return result.values.tolist()

def conv_cbc(df):
    if df is None:
        return None
    
    df_ = df.loc[df["发放银行"]=="建设银行"].copy()
    if not df_.query("提高待遇>0").empty:
        result = pd.concat([
            df_.loc[df_["企业补贴"]>0,["x_银行帐号", "姓名", "企业补贴"]].copy(),
            df_.loc[
                df_["提高待遇"]>0,["x_银行帐号", "姓名", "提高待遇"]].copy()
                .rename(columns={"提高待遇":"企业补贴"})
        ])
    else:
        result = df_.loc[df_["企业补贴"]>0,["x_银行帐号", "姓名", "企业补贴"]].copy()
    if result.empty:
        return None
    return add_index(result).values.tolist()

def conv_bocyt(df):
    if df is None:
        return None
    
    df_ = df.loc[df["发放银行"]=="中国银行_油区"].copy()
    df_["开户行"] = "中国银行"
    df_["行号"] = "41"

    if not df_.query("提高待遇>0").empty:
        result = pd.concat([
            df_.loc[df_["企业补贴"]>0,["企业补贴", "姓名", "x_银行帐号", "开户行", "行号"]].copy(),
            df_.loc[
                df_["提高待遇"]>0,["提高待遇", "姓名", "x_银行帐号", "开户行", "行号"]].copy()
                .rename(columns={"提高待遇":"企业补贴"})
        ])
    else:
        result = df_.loc[df_["企业补贴"]>0,["企业补贴", "姓名", "x_银行帐号", "开户行", "行号"]].copy()
    if result.empty:
        return None
    return result.values.tolist()

def conv_bocny(df):
    if df is None:
        return None
    df_ = df.loc[df["发放银行"]=="中国银行_南阳"].copy()
    if not df_.query("提高待遇>0").empty:
        result = pd.concat([
            df_.loc[df_["企业补贴"]>0,["姓名", "身份证", "发放银行", "x_银行帐号", "企业补贴"]].copy(),
            df_.loc[
                df_["提高待遇"]>0,["姓名", "身份证", "发放银行", "x_银行帐号"]].copy()
                .rename(columns={"提高待遇":"企业补贴"})
        ])
    else:
        result = df_.loc[df_["企业补贴"]>0,["姓名", "身份证", "发放银行", "x_银行帐号", "企业补贴"]].copy()

    if result.empty:
        return None
    return add_index(result).values.tolist()

def export_data(data,template,output):
    if data is None:
        return
    
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(template["temp_path"])
    sht = wb.sheets[template["sheet"]]    
    sht.range(template["cell"]).value = data
    wb.save(output)
    wb.close()
    app.quit()

# *** 根据数据生成拨付汇总表 ***
def group_by(df, key, agg_dict ):
    data = (
        df.query("re == 1")
            .groupby(key)
            .agg(agg_dict)
            .reset_index()
    )
    return data
    
def fill(ws: Worksheet, start_cell: str, data_list: list[list]):
    start_row = int(start_cell[1:])
    start_col = ord(start_cell[0]) - ord("A") + 1

    for i, row_data in enumerate(data_list):
        for j, cell_value in enumerate(row_data):
            ws.cell(row=start_row + i, column=start_col + j, value=cell_value)

def get_dwbm(filename: str):
    dwbm = pd.read_excel(filename)
    return dwbm

def merge_dwbm(dwbm, df, on_field, with_how="left") -> pd.DataFrame:
    #df["单位编码"] = df["dwbm"]
    result = pd.merge(dwbm, df, on=on_field, how=with_how)
    return result

def construct_data(dataframe: pd.DataFrame,type: str):
    result = {}
    match type:
        case "老人":
            #print(dataframe)
            df = group_by(dataframe,"dwbm",
            {
                "x_银行帐号": "count",
                "补贴更正": "sum",
                "误餐补贴": "sum",
                "补发补贴": "sum",
                "补发_其它": "sum",
                "其它扣款": "sum",
                "扣款_补贴": "sum"
            })
            df.rename(columns={"dwbm":"单位编码","x_银行帐号":"实发人数","补发_其它":"提高待遇"},inplace=True)
            #print(df)
            df["企业补贴"] = df["补贴更正"] + df["误餐补贴"] + df["补发补贴"]
            df["费用扣减"] = df["其它扣款"] + df["扣款_补贴"]
            df["应发金额"] = df["企业补贴"] + df["提高待遇"] 
            df["实发金额"] = df["应发金额"] - df["费用扣减"]
            dwbm = get_dwbm(r"D:\企业补贴\银行报盘\全民dwbm.xlsx")
            #print(df)
            result["单位汇总"] = merge_dwbm(dwbm,
                       df.filter(
                            items=[
                                "单位编码", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额"]), 
                        "单位编码").values.tolist()
            df = group_by(dataframe,"发放银行",
            {
                "x_银行帐号": "count",
                "补贴更正": "sum",
                "误餐补贴": "sum",
                "补发补贴": "sum",
                "补发_其它": "sum",
                "其它扣款": "sum",
                "扣款_补贴": "sum",
            })
            df.rename(
                columns={
                    "补贴更正":"补贴标准",
                    "其它扣款":"代扣大额",
                    "扣款_补贴":"补贴扣款",
                    "补发_其它":"提高待遇",
                    "x_银行帐号":"实发人数"
            },inplace=True)
            #print(df)
            df["企业补贴"] = df["补贴标准"] + df["误餐补贴"] + df["补发补贴"]
            df["费用扣减"] = df["代扣大额"] + df["补贴扣款"]
            df["应发金额"] = df["企业补贴"] + df["提高待遇"]
            df["实发金额"] = df["应发金额"] - df["费用扣减"]
            result["银行汇总"] = df.filter(
                items=["发放银行", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额"]).values.tolist()
            
            return result


        case "集体工":
            df = group_by(dataframe,"dwbm",
            {
                "x_银行帐号": "count",
                "补贴更正": "sum",
                "误餐补贴": "sum",
                "补发补贴": "sum",
                "补发_其它": "sum",
                "其它扣款": "sum",
                "扣款_补贴": "sum"
            })
            df.rename(columns={"dwbm":"单位编码","x_银行帐号":"实发人数","补发_其它":"提高待遇"},inplace=True)
            #print(df)
            df["企业补贴"] = df["补贴更正"] + df["误餐补贴"] + df["补发补贴"]
            df["费用扣减"] = df["其它扣款"] + df["扣款_补贴"]
            df["应发金额"] = df["企业补贴"] + df["提高待遇"] 
            df["实发金额"] = df["应发金额"] - df["费用扣减"]
            dwbm = get_dwbm(r"D:\企业补贴\银行报盘\非全民dwbm.xlsx")
            #print(df)
            result["企业补贴（财务拨款）"] = merge_dwbm(dwbm,
                        df.filter(
                            items=[
                                "单位编码", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额","性质"]), 
                        "单位编码").query("性质 != '工程'").filter(
                            items=[
                                "单位名称", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额","性质"]).values.tolist()
            result["企业补贴（单位征集）"] = merge_dwbm(dwbm,
                        df.filter(
                            items=[
                                "单位编码", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额","性质"]), 
                        "单位编码").query("性质 == '工程' ").filter(
                            items=[
                                "单位名称", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额","性质"]).values.tolist()
            result["企业补贴（含遗孀）"] = result["企业补贴（财务拨款）"]
            df = group_by(dataframe,"发放银行",
            {
                "x_银行帐号": "count",
                "补贴更正": "sum",
                "误餐补贴": "sum",
                "补发补贴": "sum",
                "补发_其它": "sum",
                "其它扣款": "sum",
                "扣款_补贴": "sum",
            })
            df.rename(
                columns={
                    "补贴更正":"补贴标准",
                    "其它扣款":"代扣大额",
                    "扣款_补贴":"补贴扣款",
                    "补发_其它":"提高待遇",
                    "x_银行帐号":"实发人数"
            },inplace=True)
            #print(df)
            df["企业补贴"] = df["补贴标准"] + df["误餐补贴"] + df["补发补贴"]
            df["费用扣减"] = df["代扣大额"] + df["补贴扣款"]
            df["应发金额"] = df["企业补贴"] + df["提高待遇"]
            df["实发金额"] = df["应发金额"] - df["费用扣减"]
            result["企业补贴（银行汇总）"] = df.filter(
                items=["发放银行", "实发人数", "企业补贴", "提高待遇", "费用扣减","应发金额", "实发金额"]).values.tolist()
            
            return result

        case "中人":pass
        case _: pass

def summary(data, template, type, output):
    file_name= template[type]["path"]
    #print(file_name)
    wb = load_workbook(filename=file_name)
    for sheet in template[type]["sheets"]:
        sht_name = sheet["name"]
        sht_title = sheet["title"]
        sht_arch = sheet["grid"]["arch"]

        ws = wb[sht_name]
        ws[sht_title["cell"]] = sht_title["value"]
        fill(ws, sht_arch, construct_data(data,type)[sht_name])
    wb.save(output)
    wb.close()




# *** 根据发放明细生成银行汇总表 ***




# *** 根据数据生成人员变动表 ***
def make_change_report(df, term, type, output_base):
    df['企业补贴'] = df['补贴更正']+ df['误餐补贴']
    match type:
        case "老人":
            info =(df
                   .query(
                       "re == 0 and 死亡登记 == " + f"{"'"+term[:4]+"-"+str(int(term[-2:]))+"'"}")
                   .filter(
                       items=["收款行行号", "单位名称", "姓名", "身份证", "swsj", "企业补贴"])
                       .values.tolist())
            
            title_name = f"{term[:4]+"年" + str(int(term[-2:]))+"月离退休职工（老人）人员变动汇总表"}"
            file_name = output_base / f"{"离退休职工（老人）人员变动汇总表-"+ term + ".xlsx"}"
        case "集体工":
            info = (df
                    .query(
                       "re == 0 and 死亡登记 == "  + f"{"'"+term[:4]+"-"+str(int(term[-2:]))+"'"}")
                   .filter(
                       items=["内部编码", "单位名称", "姓名", "身份证", "swsj", "企业补贴"])
                       .values.tolist())
            title_name = f"{term[:4]+"年" + str(int(term[-2:]))+"月退休集体工人员变动汇总表"}"
            file_name = output_base / f"{"退休集体工人员变动汇总表-" +term+".xlsx"}"     
        case _: 
            pass 
    #print(info)     
    report_change(arr=info,title_name=title_name,output=file_name)

def report_change(arr,  title_name, output):    
    arr_len = len(arr)
    app = xw.App(visible=False)
    wb = app.books.add()
    sht = wb.sheets[0]

    title = sht.range("A1:G2")
    title.merge()
    title.value = title_name
    title.api.font.name = "黑体"
# title.api.font.bold = True
    title.api.font.size = 18

    signature = sht["A3"]
    signature.value = "单位：社保中心养老保险室"
    signature.font.name = "楷体"
    signature.font.size = 12
    signature.api.HorizontalAlignment = -4131

    header_rng = f"A4:G4"
    header = sht[header_rng]
    header.value = [
        "序号",
        "员工编号",
        "单位",
        "姓名",
        "身份证",
        "死亡时间",
        "企业补贴金额",
    ]


    sht["A5"].options(transpose=True).value = list(range(1, arr_len + 1))
    sht[f"A5:A{arr_len+4}"].api.HorizontalAlignment = -4108

    sht[f"B5:E{arr_len+4}"].api.NumberFormat = "@"
    sht["B5"].value = arr


    sht[f"A{arr_len+5}:D{arr_len+5}"].merge()
    sht[f"A{arr_len+5}:D{arr_len+5}"].value = "合计"
    sht[f"A{arr_len+5}"].api.HorizontalAlignment = -4108

    total = sht[f"G{arr_len+5}"]
    total.formula = f"=SUM(G5:G{arr_len+4})"

    comment = sht[f"B{arr_len+7}"]
    comment.value = "*注：此表仅供参考，以实际变动为准*"
    comment.api.font.name = "仿宋"
    comment.api.font.size = 14

    # 水平居中
    sht["A1"].api.HorizontalAlignment = -4108
    sht[f"A4:G{arr_len+5}"].api.HorizontalAlignment = -4108

    # 设置列宽
    sht[f"A3:A{arr_len+4}"].column_width = 7
    sht[f"B3:B{arr_len+4}"].column_width = 11
    sht[f"C3:C{arr_len+4}"].column_width = 18
    sht[f"D3:D{arr_len+4}"].column_width = 12
    sht[f"E3:E{arr_len+4}"].column_width = 22
    sht[f"F3:F{arr_len+4}"].column_width = 10
    sht[f"G3:G{arr_len+4}"].column_width = 15

    # 设置行高
    sht["A1"].row_height = 44.5
    sht["A3"].row_height = 27.75
    sht[f"A4:F{arr_len+5}"].row_height = 21.25

    # 设置边框
    content = sht[f"A4:G{arr_len+5}"]
    content.api.BorderAround(LineStyle=1, Weight=2)

    xborder = content.api.Borders(11)
    xborder.LineStyle = 1
    xborder.Weight = 2

    yborder = content.api.Borders(12)
    yborder.LineStyle = 1
    yborder.Weight = 2


    # 设置打印区域
    print_area = f"$A$1:$G${arr_len+7}"
    sht.page_setup.print_area = print_area

    # 横向打印
    # sheet.api.PageSetup.Orientation = 2  # 1为纵向，2为横向
    # 纵向打印
    sht.api.PageSetup.Orientation = 1

    sht.api.PageSetup.Zoom = False
    sht.api.PageSetup.FitToPagesWide = 1  # 宽度调整为1页
    # sheet.api.PageSetup.FitToPagesTall = 1  # 高度调整为1页

    # 设置页边距（单位：厘米）
    # sheet.api.PageSetup.LeftMargin = 1.5
    # sheet.api.PageSetup.RightMargin = 1.5
    # sheet.api.PageSetup.TopMargin = 2
    # sheet.api.PageSetup.BottomMargin = 2

    # 水平居中
    sht.api.PageSetup.CenterHorizontally = True
    wb.save(output)
    wb.close()

    


if __name__ == '__main__':
    df_lr = read_dbf(source_path(base=source_base,term=current_term)["老人"] / "bt_ltx.dbf")
    
    #df_zr = read_dbf(source_path(base=source_base,term=current_term)["中人"] / "bt_ltx.dbf") 
    df_jtg = read_dbf(source_path(base=source_base,term=current_term)["集体工"] / "bt_ltx.dbf")
    
    #print(df_lr.query("re == 0 and 死亡登记=''" ).copy())
    export_data(
        conv_icbc(preprocess(df_lr,"老人企业补贴")),
        bank_templates["工行"],
        dest_path(dest_base,current_term)["工行"]/f"{"老人企业补贴(工行报盘)-"+current_term+".xlsx"}")
    export_data(conv_cbc(preprocess(df_lr,"老人企业补贴")),bank_templates["建行"],dest_path(dest_base,current_term)["建行"]/f"{"老人企业补贴(建行报盘)-"+current_term+".xls"}")
    #export_data(conv_bocyt(preprocess(df_lr,"老人企业补贴")),bank_templates["中行-油区"],dest_path(dest_base,current_term)["中行"]/f"{"老人企业补贴(中行-油区报盘)-"+current_term+".xlsx"}")
    export_data(conv_bocny(preprocess(df_lr,"老人企业补贴")),bank_templates["中行-南阳"],dest_path(dest_base,current_term)["中行"]/f"{"老人企业补贴(中行-南阳报盘)-"+current_term+".xlsx"}")

    export_data(conv_icbc(preprocess(df_jtg,"集体工企业补贴")),bank_templates["工行"],dest_path(dest_base,current_term)["工行"]/f"{"集体工企业补贴(工行报盘)-"+current_term+".xlsx"}")
    export_data(conv_cbc(preprocess(df_jtg,"集体工企业补贴")),bank_templates["建行"],dest_path(dest_base,current_term)["建行"]/f"{"集体工企业补贴(建行报盘)-"+current_term+".xls"}")
    export_data(conv_bocyt(preprocess(df_jtg,"集体工企业补贴")),bank_templates["中行-油区"],dest_path(dest_base,current_term)["中行"]/f"{"集体工企业补贴(中行-油区报盘)-"+current_term+".xlsx"}")
    export_data(conv_bocny(preprocess(df_jtg,"集体工企业补贴")),bank_templates["中行-南阳"],dest_path(dest_base,current_term)["中行"]/f"{"集体工企业补贴(中行-南阳报盘)-"+current_term+".xlsx"}")

    make_change_report(df_lr, current_term, "老人", Path(dest_base)/f"{current_term[:4]+"年"}" / f"{str(int(current_term[-2:]))+"月"}")
    make_change_report(df_jtg, current_term, "集体工", Path(dest_base)/f"{current_term[:4]+"年"}" / f"{str(int(current_term[-2:]))+"月"}")
    #result = construct_data(df_lr,"老人")
    out_lr = Path(dest_base)/f"{current_term[:4]+"年"}" / f"{str(int(current_term[-2:]))+"月"}"/ f"{"老人企业补贴汇总-"+ current_term+r".xlsx"}"
    out_jtg = Path(dest_base)/f"{current_term[:4]+"年"}" / f"{str(int(current_term[-2:]))+"月"}"/ f"{"集体工企业补贴汇总-"+ current_term+r".xlsx"}"
    summary(df_lr,summary_templates,'老人',str(out_lr))
    summary(df_jtg,summary_templates,'集体工',str(out_jtg))
    

    
