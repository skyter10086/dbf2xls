from openpyxl import Workbook
from . import Style, Block, PageSetup

class Report:
    def __init__(self, workbook: Workbook,name:str, page_setup:PageSetup):
        self.name = name
        self.page_setup = page_setup
        self.__wb = workbook
        self.__ws = self.__wb.create_sheet(name,0)
        self.__blocks = {}

    def add_block(self, name:str, block:Block):
        self.__blocks[name] = block

    def apply_page_setup(self):
        __page_set = self.page_setup
        self.__ws.print_area = __page_set["print_area"]
        self.__ws.print_title_rows = __page_set["title_rows"]
        self.__ws.page_setup.orientation = __page_set["orientation"]
        self.__ws.page_setup.paperSize = __page_set["paper_size"]
        self.__ws.page_margins = __page_set["page_margins"]
        self.__ws.sheet_properties.pageSetUpPr.fitToPage = __page_set["fit_page"]
        self.__ws.page_setup.fitToHeight = __page_set["fit2height"]
        self.__ws.page_setup.fitToWidth = __page_set["fit2width"]
        self.__ws.oddHeader.center.text = __page_set["header"]
        self.__ws.oddFooter.center.text = __page_set["footer"]
        self.__ws.print_options.horizontalCentered = __page_set["horizontal_centered"]
        self.__ws.print_options.verticalCentered = __page_set["vertical_centered"]

    def apply_blocks(self):
        for block_name in self.__blocks.keys():
            __block = self.__blocks[block_name]
            area = f"{__block.head}:{__block.tail}"
            __cells = self.__ws[area]
            # print(f"{block_name}(cell) : {__cells}")
            __merge = __block.merge
            __style = __block.style
            __font = __style["font"]
            __align = __style["align"]
            __border = __style["border"]
            __num_fmt = __style["num_fmt"]
            __height = __style["rows_height"]
            __width = __style["cols_width"]
            __value = __block.value

            row_head = int(__block.head[1:])
            row_tail = int(__block.tail[1:])
            col_head = __block.head[0]
            col_tail = __block.tail[0]

            # apply style
            for row in __cells:
                for i, cell in enumerate(row):
                    # print(f"{cell} is applying style:")
                    cell.font = __font
                    cell.alignment = __align
                    cell.border = __border
                    cell.number_format = __num_fmt[i]

            # apply merge
            if __merge == True:
                self.__ws.merge_cells(area)
                self.__ws[__block.head].value = __value[0][0]
            else:
                start_row = int(__block.head[1:])
                start_col = ord(__block.head[0]) - ord("A") + 1  # 列号转为数字

                for i, row_data in enumerate(__value):
                    for j, cell_value in enumerate(row_data):
                        self.__ws.cell(
                            row=start_row + i, column=start_col + j, value=cell_value
                        )

            # apply height and width
            for i, v in enumerate(list(range(row_head, row_tail + 1))):
                self.__ws.row_dimensions[v].height = __height[i]

            col_range = [chr(x) for x in range(ord(col_head), ord(col_tail) + 1)]
            col_width = dict(zip(col_range, __width))
            for col in col_range:
                self.__ws.column_dimensions[col].width = col_width.get(
                    col, 8.43
                )  # 第二个参数是默认参数

    def generate(self):
        self.apply_blocks()
        self.apply_page_setup()
