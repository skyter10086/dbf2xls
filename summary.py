from dbfread2 import DBF
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from pathlib import Path

term = "202508"
base_dir = Path("D:/","企业补贴","数据")
lr_datafile = base_dir / f"{"企业补贴"+term}" / "bt_ltx.dbf"
zr_datafile = base_dir / f"{"提高待遇"+term}" / "bt_ltx.dbf"
jtg_datafile = base_dir / f"{"集体工企业补贴"+term}" / "bt_ltx.dbf"
summary_path = Path(r"D:\企业补贴\银行报盘") / f"{term[:4]+"年"}" / f"{str(int(term[-2:]))+"月"}"

if not summary_path.exists():
    summary_path.mkdir(parents=True, exist_ok=True)

def fill(ws: Worksheet, start_cell: str, data_list: list[list]):
    start_row = int(start_cell[1:])
    start_col = ord(start_cell[0]) - ord("A") + 1

    for i, row_data in enumerate(data_list):
        for j, cell_value in enumerate(row_data):
            ws.cell(row=start_row + i, column=start_col + j, value=cell_value)


def read_dbf(
    filename: str,
    column_names: list[str],
    group_by: str,
    agg_dict: dict,
    alias: list[str],
):
    table = DBF(filepath=filename)
    df = pd.DataFrame(iter(table))
    # print(df)
    dbf = (
        df.query("RE == 1")
        .filter(items=column_names)
        .groupby(group_by)
        .agg(agg_dict)
        .reset_index()
    )
    dbf.columns = alias

    return dbf


def read_dwbm(filename: str):
    dwbm = pd.read_excel(filename)
    return dwbm


def build_data(dwbm, dbf, on_field, with_how) -> list[list]:
    result = pd.merge(dwbm, dbf, on=on_field, how=with_how)

    return result


def export_data(template: str, sheets: list[dict], output: str):
    wb = load_workbook(filename=template)

    for sheet in sheets:

        sht_name = sheet["name"]
        sht_title = sheet["title"]
        sht_arch = sheet["arch"]
        data = sheet["data"]

        ws = wb[sht_name]
        ws[sht_title["cell"]] = sht_title["value"]
        fill(ws, sht_arch, data)

    wb.save(output)


def main():
    # 老人企业补贴
    dbf = read_dbf(
        filename=str(lr_datafile),  # dbf
        column_names=[
            "DWBM",
            "X_银行帐号",
            "补贴更正",
            "误餐补贴",
            "补发补贴",
            "补发_其它",
            "其它扣款",
            "扣款_补贴",
        ],
        group_by="DWBM",
        agg_dict={
            "X_银行帐号": "count",
            "补贴更正": "sum",
            "误餐补贴": "sum",
            "补发补贴": "sum",
            "补发_其它": "sum",
            "其它扣款": "sum",
            "扣款_补贴": "sum",
        },
        alias=[
            "单位编码",
            "实发人数",
            "补贴标准",
            "误餐补贴",
            "补发补贴",
            "提高待遇",
            "代扣大额",
            "补贴扣款",
        ],
    )
    dbf["企业补贴"] = dbf["补贴标准"] + dbf["误餐补贴"] + dbf["补发补贴"]
    dbf["费用扣减"] = dbf["代扣大额"] + dbf["补贴扣款"]
    dbf["实发金额"] = dbf["企业补贴"] + dbf["提高待遇"] - dbf["费用扣减"]
    dbf_ = dbf.filter(
        items=["单位编码", "实发人数", "企业补贴", "提高待遇", "费用扣减", "实发金额"]
    )
    dwbm = read_dwbm(r"D:\企业补贴\银行报盘\全民dwbm.xlsx")
    data_qm = build_data(dwbm, dbf_, "单位编码", "left").values.tolist()
    #print(data_qm)
    dbf_1 = read_dbf(
        filename=str(lr_datafile),
        column_names=[
            "发放银行",
            "X_银行帐号",
            "补贴更正",
            "误餐补贴",
            "补发补贴",
            "补发_其它",
            "其它扣款",
            "扣款_补贴",
        ],
        group_by="发放银行",
        agg_dict={
            "X_银行帐号": "count",
            "补贴更正": "sum",
            "误餐补贴": "sum",
            "补发补贴": "sum",
            "补发_其它": "sum",
            "其它扣款": "sum",
            "扣款_补贴": "sum",
        },
        alias=[
            "发放银行",
            "实发人数",
            "补贴标准",
            "误餐补贴",
            "补发补贴",
            "提高待遇",
            "代扣大额",
            "补贴扣款",
        ],
    )
    dbf_1["企业补贴"] = dbf_1["补贴标准"] + dbf_1["误餐补贴"] + dbf_1["补发补贴"]
    dbf_1["费用扣减"] = dbf_1["代扣大额"] + dbf_1["补贴扣款"]
    dbf_1["实发金额"] = dbf_1["企业补贴"] + dbf_1["提高待遇"] - dbf_1["费用扣减"]
    dbf_1 = dbf_1.filter(
        items=["发放银行", "实发人数", "企业补贴", "提高待遇", "费用扣减", "实发金额"]
    )
    bank_data = dbf_1.values.tolist()
    export_data(
        template=r"D:\企业补贴\银行报盘\离退休职工企业补贴汇总模板（老人）.xlsx",
        sheets=[
            {
                "name": "单位汇总",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"全民离退休人员企业补贴发放汇总表(单位)"}",
                },
                "arch": "B4",
                "data": data_qm,
            },
            {
                "name": "银行汇总",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"全民离退休人员企业补贴发放汇总表(银行)"}",
                },
                "arch": "J4",
                "data": bank_data,
            },
        ],
        output= str(summary_path / f"{"老人企业补贴汇总"+term+".xlsx"}"),
    )

    # 集体工企业补贴
    dbf2 = read_dbf(
        filename=str(jtg_datafile),  # dbf
        column_names=[
            "DWBM",
            "X_银行帐号",
            "补贴更正",
            "补发补贴",
            "补发_其它",
            "其它扣款",
            "扣款_补贴",
        ],
        group_by="DWBM",
        agg_dict={
            "X_银行帐号": "count",
            "补贴更正": "sum",
            "补发补贴": "sum",
            "补发_其它": "sum",
            "其它扣款": "sum",
            "扣款_补贴": "sum",
        },
        alias=[
            "单位编码",
            "实发人数",
            "补贴标准",
            "补发补贴",
            "提高待遇",
            "代扣大额",
            "补贴扣款",
        ],
    )
    dwbm2 = read_dwbm(filename=r"D:\企业补贴\银行报盘\非全民dwbm.xlsx")
    data2 = build_data(dwbm=dwbm2, dbf=dbf2, on_field="单位编码", with_how="left")
    data2["企业补贴"] = data2["补贴标准"] + data2["补发补贴"] - data2["补贴扣款"]
    data2["代扣费用"] = data2["补贴扣款"] + data2["代扣大额"]
    data2["实发金额"] = data2["企业补贴"] + data2["提高待遇"] - data2["代扣费用"]

    data2f = data2.filter(
        items=[
            "单位名称",
            "实发人数",
            "企业补贴",
            "提高待遇",
            "代扣费用",
            "实发金额",
            "性质",
        ]
    )
    fqm = data2f[data2f["性质"] != "工程"]
    fqm_data = fqm.values.tolist()

    gc = data2f[data2f["性质"] == "工程"]
    gc_data = gc.values.tolist()

    dbf_2 = read_dbf(
        filename=str(jtg_datafile),
        column_names=[
            "发放银行",
            "X_银行帐号",
            "补贴更正",
            "补发补贴",
            "补发_其它",
            "扣款_补贴",
            "其它扣款",
        ],
        group_by="发放银行",
        agg_dict={
            "X_银行帐号": "count",
            "补贴更正": "sum",
            "补发补贴": "sum",
            "补发_其它": "sum",
            "扣款_补贴": "sum",
            "其它扣款": "sum",
        },
        alias=[
            "发放银行",
            "实发人数",
            "补贴标准",
            "补发补贴",
            "提高待遇",
            "补贴扣款",
            "代扣大额",
        ],
    )
    dbf_2["代扣费用"] = dbf_2["补贴扣款"] + dbf_2["代扣大额"]
    dbf_2["企业补贴"] = dbf_2["补贴标准"] + dbf_2["补发补贴"]
    dbf_2["实发金额"] = dbf_2["企业补贴"] + dbf_2["提高待遇"] - dbf_2["代扣费用"]

    bank_data2 = dbf_2.filter(
        items=["发放银行", "实发人数", "企业补贴", "提高待遇", "代扣费用", "实发金额"]
    ).values.tolist()

    export_data(
        template=r"D:\企业补贴\银行报盘\退休集体工企业补贴汇总模板.xlsx",
        sheets=[
            {
                "name": "企业补贴（财务拨款）",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"退休集体工企业补贴发放汇总表"}",
                },
                "arch": "B4",
                "data": fqm_data,
            },
            {
                "name": "企业补贴（含遗孀）",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(含遗孀)"}",
                },
                "arch": "B4",
                "data": fqm_data,
            },
            {
                "name": "企业补贴（单位征集）",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(工程)"}",
                },
                "arch": "K4",
                "data": gc_data,
            },
            {
                "name": "企业补贴（银行汇总）",
                "title": {
                    "cell": "A1",
                    "value": f"{term[:4]+"年"+str(int(term[-2:]))+"月"+"退休集体工企业补贴发放汇总表(银行)"}",
                },
                "arch": "I4",
                "data": bank_data2,
            },
        ],
        output=str(summary_path / f"{"退休集体工企业补贴汇总"+term+".xlsx"}") ,
    )
"""
    # 中人提高待遇
    dbf3 = read_dbf(
        filename=r"D:\企业补贴\数据\提高待遇202507\bt_ltx.dbf",  # dbf
        column_names=["DWBM", "X_银行帐号", "应发补贴", "其它扣款", "实发补贴"],
        group_by="DWBM",
        agg_dict={
            "X_银行帐号": "count",
            "应发补贴": "sum",
            "其它扣款": "sum",
            "实发补贴": "sum",
        },
        alias=["单位编码", "人数", "应发金额", "扣款", "实发金额"],
    )
    dwbm3 = read_dwbm(r"D:\企业补贴\银行报盘\中人dwbm.xlsx")
    data_zr = build_data(dwbm3, dbf3, "单位编码", "left").values.tolist()

    dbf_3 = read_dbf(
        filename=r"D:\企业补贴\数据\提高待遇202507\bt_ltx.dbf",
        column_names=["发放银行", "X_银行帐号", "应发补贴", "其它扣款", "实发补贴"],
        group_by="发放银行",
        agg_dict={
            "X_银行帐号": "count",
            "应发补贴": "sum",
            "其它扣款": "sum",
            "实发补贴": "sum",
        },
        alias=["发放银行", "人数", "应发金额", "扣款", "实发金额"],
    )
    bank_data3 = dbf_3.values.tolist()
    export_data(
        template=r"D:\企业补贴\银行报盘\离退休职工企业补贴汇总模板（中人）.xlsx",
        sheets=[
            {
                "name": "单位汇总",
                "title": {
                    "cell": "A1",
                    "value": "2025年7月退休中人提高待遇发放汇总表(单位)",
                },
                "arch": "B4",
                "data": data_zr,
            },
            {
                "name": "银行汇总",
                "title": {
                    "cell": "A1",
                    "value": "2025年7月退休中人提高待遇发放汇总表(银行)",
                },
                "arch": "H4",
                "data": bank_data3,
            },
        ],
        output=r"D:\企业补贴\银行报盘\2025年\7月\中人提高待遇汇总202507.xlsx",
    )
"""

if __name__ == "__main__":
    main()


"""
    wb = load_workbook(
        filename=r"E:\企业补贴\银行报盘\离退休职工企业补贴汇总模板（老人）.xlsx"
    )
    ws = wb["单位汇总"]
    ws["A1"] = "2025年5月全民离退休人员企业补贴发放汇总表(单位)" # title
    ws_1 = wb["银行汇总"]
    ws_1["A1"] = "2025年5月全民离退休人员企业补贴发放汇总表(银行)" # title
    
    fill(ws_1, "G4", bank_data)
    fill(ws, "B4", data)
    wb.save(r"E:\企业补贴\银行报盘\2025年\5月\老人企业补贴汇总202505.xlsx") # file
    wb2 = load_workbook(
        filename=r"E:\企业补贴\银行报盘\退休集体工企业补贴汇总模板.xlsx"
    )
    ws2 = wb2["企业补贴（财务拨款）"]
    ws2["A1"] = "2025年5月退休集体工企业补贴发放汇总表" # title
    ws_2 = wb2["企业补贴（含遗孀）"]
    ws_2["A1"] = "2025年5月退休集体工企业补贴发放汇总表(含遗孀)" # title
    ws3 = wb2["企业补贴（单位征集）"]
    ws3["A1"] = "2025年5月退休集体工企业补贴发放汇总表(工程)" # title
    fill(ws2, "B4", fqm_data)
    fill(ws_2, "B4", fqm_data)
    fill(ws3, "J4", gc_data)
    ws_2 = wb2["企业补贴（银行汇总）"]
    ws_2["A1"] = "2025年5月退休集体工企业补贴发放汇总表(银行)" # title
    fill(ws_2, "H4", bank_data2)
    wb2.save(r"E:\企业补贴\银行报盘\2025年\5月\退休集体工企业补贴汇总202505.xlsx") # file

"""
